"""
MacDirScope

Description:
    A utility for macOS to scan a directory, extract rich filesystem and
    extended metadata, and export the results into a formatted Excel file.

    The program executes the following numbered steps:
      1. Checks if the system is macOS and has the 'mdls' command-line tool.
      2. Prompts the user to select an input directory to scan via a native dialog.
      3. Prompts the user for a save location for the output Excel file.
      4. Pre-scans the directory to count items and efficiently pre-computes all
         directory sizes for a significant performance increase.
      5. Displays a progress bar and begins processing every file and folder.
      6. For each item, it extracts standard info (size, dates) and extended
         macOS metadata like Finder Tags and Kind using the 'mdls' command.
      7. Writes the collected data row-by-row into an Excel worksheet.
      8. Formats the Excel file with appropriate column widths, date formatting,
         and a frozen header row for easy viewing.
      9. Displays a final completion report summarizing the operation.

Usage:
    - Ensure required libraries are installed:
          pip install openpyxl
    - Run the script from a terminal:
          python mac-dir-scope.py

Author: Vitalii Starosta
GitHub: https://github.com/sztaroszta
License: MIT
"""

import os
import subprocess
import sys
from tkinter import Tk, filedialog, messagebox, ttk, Label, Button, Toplevel, DoubleVar
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font
from typing import Tuple, Optional, List, Dict

# --- Global Excel File Configuration ---
COLUMN_WIDTHS = {
    'A': 5,   # Row number
    'B': 25,  # Path
    'C': 11,  # Size (KB)
    'D': 19,  # Creation date
    'E': 19,  # Modified date
    'F': 10,  # Hidden status
    'G': 10,  # Tags
    'H': 15,  # Kind
    'I': 7,   # File type
}
LEVEL_COLUMN_WIDTH = 10

# --- GUI Classes ---

class ProgressWindow:
    """
    A GUI window to display the progress of a long-running task.
    It features a progress bar, a status label, and an item counter.
    """
    
    def __init__(self, total_items: int):
        """
        Initializes the ProgressWindow.

        Args:
            total_items (int): The total number of items to be processed,
                               used to scale the progress bar.
        """
        self.root = Tk()
        self.root.withdraw()
        self.progress_toplevel = Toplevel(self.root)
        
        self.window = self.progress_toplevel
        self.window.title("Processing Directory...")
        self.window.geometry("600x150")
        self.window.resizable(False, False)
        self.window.transient()
        self.window.grab_set()
        
        self.total_items = total_items
        self.setup_widgets()
        
        self.window.lift()
        self.window.attributes('-topmost', True)
    
    def setup_widgets(self):
        """Creates and arranges all the widgets within the progress window."""
        title_label = Label(self.window, text="Extracting Directory Metadata", font=('Arial', 13, 'bold'))
        title_label.pack(pady=10)
        
        self.status_label = Label(self.window, text="Initializing...")
        self.status_label.pack(pady=5)
        
        self.progress_var = DoubleVar()
        self.progress_bar = ttk.Progressbar(self.window, length=350, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(pady=10)
        
        self.progress_label = Label(self.window, text="0 / 0 items processed")
        self.progress_label.pack(pady=5)
        
        self.cancel_button = Button(self.window, text="Run in Background", command=self.minimize_window)
        self.cancel_button.pack(pady=5)

    def update_progress(self, processed: int, status: str = ""):
        """
        Updates the progress bar and status labels.

        Args:
            processed (int): The number of items processed so far.
            status (str, optional): A message describing the current operation.
        """
        progress_percent = (processed / self.total_items * 100) if self.total_items > 0 else 0
        self.progress_var.set(progress_percent)
        if status:
            self.status_label.config(text=status)
        self.progress_label.config(text=f"{processed} / {self.total_items} items processed")
        self.window.update()
    
    def minimize_window(self):
        """Minimizes the progress window to the dock."""
        self.window.iconify()
    
    def close(self):
        """Destroys the progress window and its Tkinter root."""
        try:
            self.root.destroy()
        except:
            pass

class CompletionReportWindow:
    """
    A GUI dialog that displays a summary of the processing results.

    This window runs its own mainloop to act as a blocking dialog, waiting for
    user confirmation before the script fully exits.
    """
    
    def __init__(self, stats: dict):
        """
        Initializes and displays the completion report.

        Args:
            stats (dict): A dictionary containing statistics from the operation.
        """
        self.window = Tk()
        self.window.title("Processing Complete")
        self.window.geometry("550x350")
        self.window.resizable(False, False)
        
        self.stats = stats
        self.setup_widgets()
        
        self.window.lift()
        self.window.attributes('-topmost', True)
        
        self.window.mainloop()

    def setup_widgets(self):
        """Creates and arranges all widgets within the report window."""
        title_label = Label(self.window, text="✓ Processing Complete", font=('Arial', 15, 'bold'), fg='green')
        title_label.pack(pady=15)
        
        stats_frame = ttk.Frame(self.window)
        stats_frame.pack(pady=10, padx=20, fill='both', expand=True)
        
        stats_text = f"""Directory Metadata Extraction Results:

Directory Scanned: {self.stats.get('directory', 'N/A')}
Items Processed: {self.stats.get('processed_items', 0):,}
   • Directories: {self.stats.get('directories', 0):,}
   • Files: {self.stats.get('files', 0):,}
Max Depth: {self.stats.get('max_levels', 0)} levels
Total Size: {self.stats.get('total_size_mb', 0):.2f} MB
Output File: {self.stats.get('output_file', 'N/A')}
Processing Time: {self.stats.get('duration', 'N/A')}"""
        
        stats_label = Label(stats_frame, text=stats_text, justify='left', font=('Courier', 12))
        stats_label.pack(pady=10)
        
        button_frame = ttk.Frame(self.window)
        button_frame.pack(pady=15)
        
        close_button = Button(button_frame, text="Close", command=self.window.destroy, width=15)
        close_button.pack(side='right', padx=5)
        
        if self.stats.get('output_file'):
            open_button = Button(button_frame, text="Open File Location", command=self.open_file_location, width=15)
            open_button.pack(side='left', padx=5)
    
    def open_file_location(self):
        """Opens the output file's location in the system's file explorer."""
        try:
            output_file = self.stats.get('output_file')
            if output_file and os.path.exists(output_file):
                subprocess.run(['open', '-R', output_file])
        except Exception as e:
            print(f"Could not open file location: {e}")

# --- Metadata Extraction Functions ---

def check_mdls_availability() -> bool:
    """
    Checks if the macOS 'mdls' command-line tool is available.

    Returns:
        bool: True if mdls is found and executable, otherwise False.
    """
    try:
        subprocess.run(['mdls', '--help'], capture_output=True, check=True)
        return True
    except (subprocess.CalledProcessError, FileNotFoundError):
        return False

def get_file_tags(path: str) -> str:
    """
    Retrieves Finder tags for a given file or folder using 'mdls'.

    Args:
        path (str): The full path to the file or folder.

    Returns:
        str: A comma-separated string of tags, or an empty string if none exist.
    """
    try:
        result = subprocess.run(
            ['mdls', '-name', 'kMDItemUserTags', '-raw', path],
            capture_output=True, text=True, timeout=10
        )
        return process_tags(result.stdout.strip())
    except:
        return ""

def process_tags(tags_str: str) -> str:
    """
    Cleans the raw tag output from the 'mdls' command.

    Args:
        tags_str (str): The raw string output from the mdls command.

    Returns:
        str: A clean, comma-separated string of tags.
    """
    if not tags_str or tags_str == "(null)": return ""
    tags_str = tags_str.strip('()')
    return ', '.join([tag.strip().strip('"') for tag in tags_str.split(',') if tag.strip()]) if tags_str else ""

def get_file_kind(path: str) -> str:
    """
    Retrieves the 'Kind' metadata for a file or folder (e.g., "PDF Document").

    Args:
        path (str): The full path to the file or folder.

    Returns:
        str: The Kind description, or an empty string if not available.
    """
    try:
        result = subprocess.run(
            ['mdls', '-name', 'kMDItemKind', '-raw', path],
            capture_output=True, text=True, timeout=10
        )
        kind = result.stdout.strip()
        return kind if kind != "(null)" else ""
    except:
        return ""

# --- Filesystem and Excel Processing Functions ---

def precompute_directory_sizes(root_path: str) -> Dict[str, int]:
    """
    Performs a single walk of the directory tree to calculate the total size
    of every subdirectory. This is far more efficient than re-calculating
    for each directory individually.

    Args:
        root_path (str): The top-level directory to start the scan from.

    Returns:
        Dict[str, int]: A dictionary mapping each directory's full path to its
                        total size in bytes.
    """
    dir_sizes = {}
    for root, _, files in os.walk(root_path):
        try:
            size = sum(os.path.getsize(os.path.join(root, f)) for f in files if not os.path.islink(os.path.join(root, f)))
            dir_sizes[root] = size
        except OSError:
            # Ignore directories we can't access
            dir_sizes[root] = 0
            
    # Aggregate sizes up the directory tree from deepest to shallowest
    for path in sorted(dir_sizes.keys(), key=len, reverse=True):
        parent = os.path.dirname(path)
        if parent != path and parent in dir_sizes:
            dir_sizes[parent] += dir_sizes[path]
            
    return dir_sizes

def get_file_info(path: str, directory_sizes: Dict[str, int]) -> Optional[Tuple]:
    """
    Gathers all standard and extended metadata for a single file or folder.

    Args:
        path (str): The full path to the item.
        directory_sizes (Dict[str, int]): A pre-computed dictionary of directory sizes.

    Returns:
        Optional[Tuple]: A tuple containing all metadata fields, or None if an
                         error occurs.
    """
    try:
        stat_info = os.stat(path)
        created = datetime.fromtimestamp(stat_info.st_birthtime)
        modified = datetime.fromtimestamp(stat_info.st_mtime)
        
        if os.path.isdir(path):
            # Fast lookup from pre-computed dictionary
            size_in_bytes = directory_sizes.get(path, 0)
            size = size_in_bytes / 1024
            file_type = "Folder"
        else:
            size = stat_info.st_size / 1024
            _, ext = os.path.splitext(os.path.basename(path))
            file_type = ext[1:] if ext else "File"
        
        basename = os.path.basename(path)
        hidden = "hidden" if basename.startswith('.') else "temporary" if basename.startswith('~$') else "visible"
        
        return (created, modified, size, file_type, hidden, get_file_tags(path), get_file_kind(path))
    except:
        return None

def get_path_levels(path: str) -> List[str]:
    """
    Splits a full file path into its constituent directory levels.

    Args:
        path (str): The full file path.

    Returns:
        List[str]: A list where each element is a directory in the path.
    """
    return [level for level in path.split(os.sep) if level]

def count_files_and_max_levels(starting_directory: str) -> Tuple[int, int]:
    """
    Performs a pre-scan of a directory to get the total item count and max depth.

    Args:
        starting_directory (str): The path to the directory to scan.

    Returns:
        Tuple[int, int]: A tuple containing the total number of items and the
                         maximum directory depth found.
    """
    total_items, max_levels = 0, 0
    try:
        for root, dirs, files in os.walk(starting_directory):
            for item in dirs + files:
                total_items += 1
                path = os.path.join(root, item)
                max_levels = max(max_levels, len(get_path_levels(path)))
    except:
        pass
    return total_items, max_levels

def setup_worksheet_headers(worksheet, max_levels: int) -> List[str]:
    """
    Writes the header row to the Excel worksheet.

    Args:
        worksheet: The openpyxl worksheet object.
        max_levels (int): The number of 'Level' columns to create.

    Returns:
        List[str]: The list of header titles.
    """
    headers = ['#', 'Path', 'Size (KB)', 'Creation Date', 'Last Modified', 'Is Hidden?', 'Tags', 'Kind', 'File Type']
    headers.extend([f'Level {i+1}' for i in range(max_levels)])
    worksheet.append(headers)
    return headers

def format_worksheet(worksheet, headers: List[str]):
    """
    Applies column widths, date formatting, and freezes the top row.

    Args:
        worksheet: The openpyxl worksheet object.
        headers (List[str]): The list of header titles.
    """
    date_style = NamedStyle(name='datetime', number_format='YYYY-MM-DD HH:MM:SS')
    for col_letter, width in COLUMN_WIDTHS.items():
        worksheet.column_dimensions[col_letter].width = width
    for col in worksheet.columns:
        if col[0].value and 'Level' in str(col[0].value):
            worksheet.column_dimensions[col[0].column_letter].width = LEVEL_COLUMN_WIDTH
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    worksheet.freeze_panes = 'C2'
    worksheet.auto_filter.ref = worksheet.dimensions
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=4, max_col=5):
        for cell in row:
            cell.style = date_style
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = '0.00'

def generate_excel(starting_directory: str, save_path: str) -> Tuple[bool, dict]:
    """
    The main worker function to orchestrate the entire scanning and export process.

    Args:
        starting_directory (str): The directory to scan.
        save_path (str): The file path for the output Excel file.

    Returns:
        Tuple[bool, dict]: A tuple containing a success flag and a dictionary
                           of final processing statistics.
    """
    start_time = datetime.now()
    total_items, max_levels = count_files_and_max_levels(starting_directory)
    if total_items == 0:
        return False, {}
    
    # --- Performance Optimization ---
    print("Pre-computing directory sizes for performance...")
    directory_sizes = precompute_directory_sizes(starting_directory)
    print("Pre-computation complete. Starting main processing...")
    
    progress_window = ProgressWindow(total_items)
    progress_window.update_progress(0, "Setting up...")
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Directory Info'
    headers = setup_worksheet_headers(worksheet, max_levels)
    
    stats = {
        'directory': starting_directory, 'output_file': save_path,
        'total_items': total_items, 'max_levels': max_levels,
        'processed_items': 0, 'directories': 0, 'files': 0, 'errors': 0,
        'total_size_mb': 0, 'duration': '0s'
    }
    
    row_number, processed_items, total_size_bytes = 1, 0, 0
    try:
        for root, dirs, files in os.walk(starting_directory):
            all_items = [(d, True) for d in dirs] + [(f, False) for f in files]
            for item_name, is_dir in all_items:
                current_path = os.path.join(root, item_name)
                # Pass the pre-computed dictionary to the function
                file_info = get_file_info(current_path, directory_sizes)
                
                if file_info:
                    created, mod, size, ftype, hidden, tags, kind = file_info
                    path_levels = get_path_levels(current_path)
                    row_data = [row_number, current_path, size, created, mod, hidden, tags, kind, ftype, *path_levels]
                    worksheet.append(row_data)
                    row_number += 1
                    stats['directories' if is_dir else 'files'] += 1
                else:
                    stats['errors'] += 1
                
                processed_items += 1
                stats['processed_items'] = processed_items
                if processed_items % 10 == 0:
                    progress_window.update_progress(processed_items, f"Processing: {os.path.basename(current_path)}")

    except Exception as e:
        print(f"An error occurred during file processing: {e}")
        stats['errors'] += 1
        progress_window.close()
        return False, stats
    
    # Manually set the total size from the pre-computed value for the root directory
    stats['total_size_mb'] = directory_sizes.get(starting_directory, 0) / (1024 * 1024)
    
    progress_window.update_progress(processed_items, "Formatting and saving...")
    format_worksheet(worksheet, headers)
    
    try:
        workbook.save(save_path)
        stats['duration'] = str(datetime.now() - start_time).split('.')[0]
        progress_window.close()
        return True, stats
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        progress_window.close()
        return False, stats

# --- Main Application Logic ---

def get_directory_and_save_path() -> Tuple[Optional[str], Optional[str]]:
    """
    Prompts the user for input/output paths using the 'Create-Use-Destroy' pattern.

    It creates a temporary Tk root for each dialog to ensure it appears in front.

    Returns:
        Tuple[Optional[str], Optional[str]]: A tuple of (directory_path, save_path),
                                             or (None, None) if cancelled.
    """
    root_dir = Tk()
    root_dir.withdraw()
    starting_directory = filedialog.askdirectory(title="Select the directory to scan")
    root_dir.destroy()
    if not starting_directory:
        return None, None
    
    root_save = Tk()
    root_save.withdraw()
    directory_name = os.path.basename(starting_directory)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    default_filename = f"{directory_name}_{timestamp}.xlsx"
    save_path = filedialog.asksaveasfilename(
        title="Save Excel file as...", initialfile=default_filename,
        defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    root_save.destroy()
    if not save_path:
        return None, None
    
    return starting_directory, save_path

def main():
    """The main entry point for the script."""
    print("macOS Directory Metadata Extractor")
    print("=" * 40)
    
    if not check_mdls_availability():
        root_err = Tk()
        root_err.withdraw()
        messagebox.showerror("Dependency Error", "This script requires macOS and the 'mdls' command.")
        root_err.destroy()
        sys.exit(1)

    starting_directory, save_path = get_directory_and_save_path()
    
    if not starting_directory or not save_path:
        print("Operation cancelled by user.")
        sys.exit(0)
    
    print(f"Scanning directory: {starting_directory}")
    print(f"Output file: {save_path}")
    
    success, stats = generate_excel(starting_directory, save_path)
    
    if success:
        print("\nOperation completed successfully!")
        CompletionReportWindow(stats)
    else:
        print("\nOperation failed. Please check the error messages above.")
        root_err = Tk()
        root_err.withdraw()
        messagebox.showerror("Error", "Processing failed. Please check the console for details.")
        root_err.destroy()
        sys.exit(1)

# Entry point for the script.
if __name__ == "__main__":
    main()